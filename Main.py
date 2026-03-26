import os
import sys
import cv2 as cv
import numpy as np
import openpyxl as op

from ui.MainWindow import Ui_MainWindow
from PyQt5.QtWidgets import (QMainWindow, QApplication, QFileDialog, QMessageBox,
                              QGraphicsScene, QGraphicsPixmapItem, QGraphicsView)
from PyQt5.QtGui import QPixmap, QImage, QIcon
from PyQt5.QtCore import Qt, QThread
from PyQt5 import QtCore, QtGui
from DLSeg import DLsegBatch
from qt_material import apply_stylesheet
from ModelTrain import train_model
import json
from RootTraitCal import RootTraitCal
from ImageProcess import *


class SegThread(QThread):
    signal_img1_set = QtCore.pyqtSignal(np.ndarray)
    signal_img2_set = QtCore.pyqtSignal(np.ndarray)
    signal_process_set = QtCore.pyqtSignal(int)

    def __init__(self, filedirpath, weightpath, savepath):
        super(SegThread, self).__init__()
        self.filedirpath = filedirpath
        self.weightpath = weightpath
        self.savepath = savepath

    def __del__(self):
        self.wait()

    def run(self):
        file = os.listdir(self.filedirpath)
        for i, name in enumerate(file, start=1):
            filepath = os.path.join(self.filedirpath, name)
            img = DLsegBatch(filepath, self.weightpath, name, self.savepath)
            img2 = cv.imread(filepath)
            processnumber = int(i / len(file) * 100)
            img = img_resize(img, 1024)
            img2 = img_resize(img2, 1024)

            self.signal_img1_set.emit(img)
            self.signal_img2_set.emit(img2)
            self.signal_process_set.emit(processnumber)

            print(f'{name} has been successfully segmented')

        print('All is done')


class TraitsThread(QThread):
    signal_result_set = QtCore.pyqtSignal(np.ndarray)
    signal_result2_set = QtCore.pyqtSignal(np.ndarray)
    signal_process2_set = QtCore.pyqtSignal(int)

    def __init__(self, filedirpath, savepath, DS_pericycle, DS2_pericycle, Area_pericycle, DS1_endodermis,
                 DS2_endodermis, Area_endodermis, DS1_exodermis, DS2_exodermis, Area_exodermis):
        super(TraitsThread, self).__init__()
        self.filedirpath = filedirpath
        self.savepath = savepath
        self.DS_pericycle = DS_pericycle
        self.DS2_pericycle = DS2_pericycle
        self.Area_pericycle = Area_pericycle
        self.DS1_endodermis = DS1_endodermis
        self.DS2_endodermis = DS2_endodermis
        self.Area_endodermis = Area_endodermis
        self.DS1_exodermis = DS1_exodermis
        self.DS2_exodermis = DS2_exodermis
        self.Area_exodermis = Area_exodermis

    def __del__(self):
        self.wait()

    def run(self):
        wb = op.Workbook()
        ws_trait = wb.create_sheet('traits')
        ws_stele = wb.create_sheet('stele')
        ws_mexylem = wb.create_sheet('mexylem')
        ws_pericycle = wb.create_sheet('pericycle')
        ws_endodermis = wb.create_sheet('endodermis')
        ws_cortex = wb.create_sheet('cortex')
        ws_epidermis = wb.create_sheet('epidermis')

        traits_name_index = ['name']

        path_seg = os.path.join(self.filedirpath, 'in')
        file = os.listdir(path_seg)

        if not os.path.exists(self.savepath):
            os.makedirs(self.savepath)

        for i, name in enumerate(file, start=1):
            traits_all_index = [name]
            img_in_path = os.path.join(self.filedirpath, 'in', name)
            img_out_path = os.path.join(self.filedirpath, 'out', name)
            img_in = cv.imread(img_in_path, 0)
            img_out = cv.imread(img_out_path, 0)
            img_all = cv.add(img_in, img_out)
            processnumber2 = int(i / len(file) * 100)
            RootTrait = RootTraitCal(name, img_in, img_out, self.DS_pericycle, self.DS2_pericycle,
                                     self.Area_pericycle, self.DS1_endodermis, self.DS2_endodermis,
                                     self.Area_endodermis, self.DS1_exodermis, self.DS2_exodermis,
                                     self.Area_exodermis)

            try:
                cell_annotation, cell_area, trait_name, trait, img_last = RootTrait.trait_all_flow()
                traits_name_index.extend(trait_name)

                if i == 1:
                    ws_trait.append(traits_name_index)

                traits_all_index.extend(trait)
                ws_trait.append(traits_all_index)

                area_pericyle = [name]
                area_mexylem = [name]
                area_stele = [name]
                area_endodermis = [name]
                area_epidermis = [name]
                area_cortex = [name]

                for area_index in range(len(cell_area['area'])):
                    area_entry = cell_area['area'][area_index]
                    if 'pericycle' in area_entry:
                        area_pericyle.append(area_entry['pericycle'])
                    elif 'mexylem' in area_entry:
                        area_mexylem.append(area_entry['mexylem'])
                    elif 'stele' in area_entry:
                        area_stele.append(area_entry['stele'])
                    elif 'endodermis' in area_entry:
                        area_endodermis.append(area_entry['endodermis'])
                    elif 'epidermis' in area_entry:
                        area_epidermis.append(area_entry['epidermis'])
                    elif 'cortex' in area_entry:
                        area_cortex.append(area_entry['cortex'])

                ws_pericycle.append(area_pericyle)
                ws_mexylem.append(area_mexylem)
                ws_stele.append(area_stele)
                ws_endodermis.append(area_endodermis)
                ws_epidermis.append(area_epidermis)
                ws_cortex.append(area_cortex)

                if not os.path.exists(os.path.join(self.savepath, 'ImgResult')):
                    os.makedirs(os.path.join(self.savepath, 'ImgResult'))

                wb.save(os.path.join(self.savepath, 'result.xlsx'))
                cv.imwrite(os.path.join(self.savepath, 'ImgResult', name), img_last)
                print(f'{name} is done')

                img_last = img_resize(img_last, 1024)
                self.signal_result_set.emit(img_last)

            except Exception as e:
                print(f'{name} something went wrong: {e}')

            img_all = img_resize(img_all, 1024)
            self.signal_result2_set.emit(img_all)
            self.signal_process2_set.emit(processnumber2 + 1)

        print('All is done')


class TrainThread(QThread):
    def __init__(self, datapath, savepath, LR, ITERS, BS, SI, LI, CLA):
        super(TrainThread, self).__init__()
        self.datapath = datapath
        self.savepath = savepath
        self.LR = LR
        self.ITERS = ITERS
        self.BS = BS
        self.SI = SI
        self.LI = LI
        self.CLA = CLA

    def __del__(self):
        self.wait()

    def run(self):
        train_model(self.datapath, self.savepath, self.LR, self.ITERS, self.BS, self.SI, self.LI, self.CLA)


class EmittingStr(QtCore.QObject):
    textWritten = QtCore.pyqtSignal(str)

    def write(self, text):
        self.textWritten.emit(str(text))


class UIMain(QMainWindow):

    VALID_IMAGE_EXTENSIONS = ('.bmp', '.dib', '.png', '.jpg', '.jpeg',
                              '.pbm', '.pgm', '.ppm', '.tif', '.tiff')

    def __init__(self):
        super().__init__()
        self.CellUI = Ui_MainWindow()
        self.CellUI.setupUi(self)
        self.setWindowIcon(QIcon('logo.png'))
        self.imgtest = None
        self.imggray = None
        self.weightpath = None
        self.filedirpath = None
        self.filedirpath2 = None
        self.savepath = 'output/'
        self.Trainpath = 'data/'
        self.modelsavepath = 'output/'
        self.DLsavepath = 'output/'

        self.CellUI.pushButton_52.clicked.connect(self.open_weight)
        self.CellUI.pushButton_51.clicked.connect(self.deeplearning_seg)
        self.CellUI.pushButton_24.clicked.connect(
            lambda: self._choose_directory('filedirpath', 'The filedir path is '))
        self.CellUI.pushButton_54.clicked.connect(
            lambda: self._choose_directory('filedirpath2', 'The segmented image filedir path is '))
        self.CellUI.pushButton_55.clicked.connect(
            lambda: self._choose_directory('savepath', 'The save path is '))
        self.CellUI.pushButton_57.clicked.connect(self.get_result)
        self.CellUI.pushButton_26.clicked.connect(
            lambda: self._choose_directory('Trainpath', 'The data path is '))
        self.CellUI.pushButton_27.clicked.connect(
            lambda: self._choose_directory('modelsavepath', 'The model save path is '))
        self.CellUI.pushButton_25.clicked.connect(self._choose_dl_save_path)
        self.CellUI.pushButton_28.clicked.connect(self.train)

        sys.stdout = EmittingStr(textWritten=self.outputWritten)
        sys.stdin = EmittingStr(textWritten=self.outputWritten)

        self.CellUI.lineEdit.setText('0.1')
        self.CellUI.lineEdit_2.setText('20000')
        self.CellUI.lineEdit_3.setText('1')
        self.CellUI.lineEdit_4.setText('100')
        self.CellUI.lineEdit_5.setText('10')
        self.CellUI.lineEdit_6.setText('3')

    def outputWritten(self, text):
        cursor = self.CellUI.textBrowser.textCursor()
        cursor.movePosition(QtGui.QTextCursor.End)
        cursor.insertText(text)
        self.CellUI.textBrowser.setTextCursor(cursor)
        self.CellUI.textBrowser.ensureCursorVisible()

    def process_bar1(self, pv):
        self.CellUI.progressBar.setMinimum(0)
        self.CellUI.progressBar.setMaximum(100)
        self.CellUI.progressBar.setValue(pv)

    def _numpy_to_pixmap(self, image):
        """Convert a numpy array to a QPixmap, normalizing the color format first."""
        if len(image.shape) == 2:
            image = cv.cvtColor(image, cv.COLOR_GRAY2RGB)
        elif image.shape[2] == 4:
            image = cv.cvtColor(image, cv.COLOR_BGRA2RGB)
        elif image.shape[2] == 1:
            image = cv.cvtColor(image, cv.COLOR_GRAY2RGB)
        else:
            image = cv.cvtColor(image, cv.COLOR_BGR2RGB)

        if len(image.shape) == 3 and image.shape[2] == 3:
            return QPixmap(QImage(image.data, image.shape[1], image.shape[0], QImage.Format_RGB888))
        else:
            return QPixmap(QImage(image.data, image.shape[1], image.shape[0], QImage.Format_Grayscale8))

    def _show_image_on_view(self, image, graphics_view):
        """Display a numpy image on the given QGraphicsView widget."""
        if isinstance(image, np.ndarray):
            pixmap = self._numpy_to_pixmap(image)
        elif isinstance(image, str):
            pixmap = QPixmap(image)
        elif isinstance(image, QPixmap):
            pixmap = image
        else:
            raise TypeError(f'Unsupported image type: {type(image)}')

        pixmap = pixmap.scaled(graphics_view.width() - 3, graphics_view.height() - 3,
                               Qt.KeepAspectRatio)
        graph_scene = QGraphicsScene()
        graph_scene.addPixmap(pixmap)
        graph_scene.update()
        graphics_view.setScene(graph_scene)

    def show_img(self, image):
        self._show_image_on_view(image, self.CellUI.graphicsView_2)

    def show_img2(self, image):
        self._show_image_on_view(image, self.CellUI.graphicsView)

    def warning(self):
        title = 'Warning'
        info = 'Please select a valid file'
        message1 = QMessageBox.warning(self, title, info, QMessageBox.Yes | QMessageBox.Cancel)
        return message1

    def open_img(self):
        m = QFileDialog.getOpenFileName(None, 'Select Image File', '.')
        filepath = m[0]
        while not filepath.lower().endswith(self.VALID_IMAGE_EXTENSIONS):
            message1 = self.warning()
            if message1 == 4194304:
                return
            m = QFileDialog.getOpenFileName(None, 'Select Image File', '.')
            filepath = m[0]
        print(f'The image filepath is {filepath}')
        self.imgtest = cv.imread(filepath)
        self.show_img(self.imgtest)

    def open_weight(self):
        m = QFileDialog.getOpenFileName(None, 'Select Weight File', '.')
        filepath = m[0]
        while not filepath.lower().endswith('.pdparams'):
            message1 = self.warning()
            if message1 == 4194304:
                return
            m = QFileDialog.getOpenFileName(None, 'Select Weight File', '.')
            filepath = m[0]
        self.weightpath = filepath
        print(f'The weight path is {filepath}')

    def _choose_directory(self, attr_name, log_prefix):
        """Generic directory chooser that sets a named attribute and logs the selection."""
        path = QFileDialog.getExistingDirectory(None, 'Select Folder', '.')
        if path:
            setattr(self, attr_name, path)
            print(f'{log_prefix}{path}')

    def _choose_dl_save_path(self):
        path = QFileDialog.getExistingDirectory(None, 'Select Folder', '.')
        if path:
            self.DLsavepath = os.path.join(path, 'output')
            print(f'The predicted image save path is {path}')

    def deeplearning_seg(self):
        print('Begin to seg')
        self.thread = SegThread(self.filedirpath, self.weightpath, self.DLsavepath)
        self.thread.signal_img1_set.connect(self.show_img2)
        self.thread.signal_img2_set.connect(self.show_img)
        self.thread.signal_process_set.connect(self.process_bar1)
        self.thread.start()

    def get_result(self):
        print('Begin to calculate traits')
        try:
            DS_pericycle = float(self.CellUI.lineEdit_8.text())
            DS2_pericycle = float(self.CellUI.lineEdit_9.text())
            Area_pericycle = float(self.CellUI.lineEdit_10.text())
            DS1_endodermis = float(self.CellUI.lineEdit_11.text())
            DS2_endodermis = float(self.CellUI.lineEdit_12.text())
            Area_endodermis = float(self.CellUI.lineEdit_13.text())
            DS1_exodermis = float(self.CellUI.lineEdit_14.text())
            DS2_exodermis = float(self.CellUI.lineEdit_15.text())
            Area_exodermis = float(self.CellUI.lineEdit_16.text())
        except ValueError:
            QMessageBox.warning(self, 'Invalid Input',
                                'Please enter valid numeric values in all trait parameter fields.')
            return

        self.thread = TraitsThread(self.filedirpath2, self.savepath,
                                   DS_pericycle, DS2_pericycle, Area_pericycle, DS1_endodermis,
                                   DS2_endodermis, Area_endodermis, DS1_exodermis, DS2_exodermis,
                                   Area_exodermis)
        self.thread.signal_result_set.connect(self.show_img2)
        self.thread.signal_result2_set.connect(self.show_img)
        self.thread.signal_process2_set.connect(self.process_bar1)
        self.thread.start()

    def train(self):
        LR = self.CellUI.lineEdit.text()
        ITERS = self.CellUI.lineEdit_2.text()
        BS = self.CellUI.lineEdit_3.text()
        SI = self.CellUI.lineEdit_4.text()
        LI = self.CellUI.lineEdit_5.text()
        CLA = self.CellUI.lineEdit_6.text()

        print('Begin to train model')
        self.thread = TrainThread(self.Trainpath, self.modelsavepath, LR, ITERS, BS, SI, LI, CLA)
        self.thread.start()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    cell_ui = UIMain()
    apply_stylesheet(app, theme='dark_blue.xml')
    cell_ui.show()
    sys.exit(app.exec_())
